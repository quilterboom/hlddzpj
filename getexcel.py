# 读取和写入Excel
from openpyxl import load_workbook

class Excel:
    def __init__(self,path,sheet=None):
        self.path = path
        self.sheet = sheet
        self.list_value = []

    def get_excel(self):
        wb = load_workbook(self.path)
        if self.sheet is None:
            ws = wb.active
        else:
            ws = wb[self.sheet]
        value_k = tuple(ws.iter_rows(max_row=1,values_only=True))
        value_k1 = value_k[0]
        for i in tuple(ws.iter_rows(min_row=2,values_only=True)):
            a = dict(zip(value_k1,i))
            self.list_value.append(a)
        return self.list_value

    def write_excel(self,row,colum,value):
        wb = load_workbook(self.path)
        if self.sheet is None:
            ws = wb.active
        else:
            ws = wb[self.sheet]
        ws.cell(row,colum).value = value
        wb.save(self.path)
        wb.close()


if __name__ == '__main__':
    path = r'C:\Users\v_yupengliu\Desktop\checkdata.xlsx'